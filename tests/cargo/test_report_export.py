from pathlib import Path

from openpyxl import load_workbook

from document_assistant.ai.preprocessor import ProcessingTask
from document_assistant.cargo.matrix_postprocessor import CandidateBatch, RawClause
from document_assistant.cargo.models import ReconciliationReport, ReconciliationRow
from document_assistant.cargo.report_export import CandidateReportExport, CargoReportExport
from document_assistant.cargo.reconciliation_writer import ReconciliationExcelWriter


class TestCandidateReportExport:
    def test_returns_batch_rows_as_candidates(self):
        task = ProcessingTask(request_id=0, file_path="policy.docx")
        export = CandidateReportExport(task)
        batch = CandidateBatch(rows=[RawClause(clause_id="3.2", effective_text="Оборудование")])

        result = export.response(batch)

        assert result == {"candidates": batch.rows}

    def test_stores_task_for_debug_json_cache(self):
        task = ProcessingTask(request_id=0, file_path="policy.docx")
        export = CandidateReportExport(task)
        assert export._task is task


class TestCargoReportExport:
    def test_writes_sibling_file_and_returns_dict(self, tmp_path: Path):
        decl_path = tmp_path / "200.xlsx"
        decl_path.write_text("placeholder", encoding="utf-8")
        task = ProcessingTask(request_id=1, file_path=str(decl_path), user_name="Иванов И.И.")

        export = CargoReportExport(task, declaration_number="200", writer=ReconciliationExcelWriter())
        report = ReconciliationReport(
            declaration_number="200",
            rows=[ReconciliationRow("200", "Объект страхования", "3.2", "совпадает", "ОК")],
        )

        result = export.response(report)

        assert result["request_id"] == 1
        assert result["user_name"] == "Иванов И.И."
        assert result["declaration_number"] == "200"
        assert result["row_count"] == 1
        assert result["warnings"] == []

        output_path = Path(result["output_file"])
        assert output_path.name == "200 – результат проверки.xlsx"
        assert output_path.parent == tmp_path
        assert output_path.exists()

        wb = load_workbook(output_path)
        assert wb.active.cell(row=2, column=1).value == "200"
