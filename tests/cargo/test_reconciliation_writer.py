from pathlib import Path

from openpyxl import load_workbook

from document_assistant.cargo.models import ReconciliationReport, ReconciliationRow
from document_assistant.cargo.reconciliation_writer import ReconciliationExcelWriter
from document_assistant.cargo.response_template import ResponseTemplate, TemplateField
from document_assistant.reports.writers import ReportWriter

TEMPLATES_DIR = Path(__file__).parents[2] / "document_assistant" / "cargo" / "templates"
VERTICAL_TEMPLATE = TEMPLATES_DIR / "reconciliation_form_vertical.xlsx"
HORIZONTAL_TEMPLATE = TEMPLATES_DIR / "reconciliation_form_horizontal.xlsx"


def _sample_report() -> ReconciliationReport:
    return ReconciliationReport(
        declaration_number="200",
        rows=[
            ReconciliationRow("200/1", "Объект страхования", "3.2", "совпадает", "Полное совпадение"),
            ReconciliationRow("200/1", "Маршрут", "4.1", "не совпадает", "Выходит за территорию"),
            ReconciliationRow("200/2", "Объект страхования", "3.2", "не совпадает", "Пункт не найден"),
        ],
    )


def _stub_template(tmp_path: Path, fields: list[TemplateField]) -> ResponseTemplate:
    """A template whose path doesn't exist forces the fresh-workbook path
    while still exercising the prescribed-field skeleton."""
    return ResponseTemplate(path=tmp_path / "missing.xlsx", fields=fields)


class TestWithoutTemplate:
    """No template — the writer dumps exactly the rows the model returned."""

    def test_extends_shared_report_writer_abc(self):
        assert isinstance(ReconciliationExcelWriter(), ReportWriter)

    def test_falls_back_to_fresh_workbook(self, tmp_path: Path):
        writer = ReconciliationExcelWriter()
        output_path = tmp_path / "out.xlsx"

        writer.write(_sample_report(), output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Номер декларации/строки"
        assert ws.cell(row=2, column=1).value == "200/1"
        assert ws.cell(row=4, column=1).value == "200/2"

    def test_header_row_is_frozen(self, tmp_path: Path):
        writer = ReconciliationExcelWriter()
        output_path = tmp_path / "out.xlsx"
        writer.write(_sample_report(), output_path)
        assert load_workbook(output_path).active.freeze_panes == "A2"

    def test_column_widths_autofit_to_content(self, tmp_path: Path):
        writer = ReconciliationExcelWriter()
        output_path = tmp_path / "out.xlsx"
        writer.write(_sample_report(), output_path)
        ws = load_workbook(output_path).active
        assert ws.column_dimensions["A"].width == len("Номер декларации/строки") + 2

    def test_special_conditions_sheet_added_when_present(self, tmp_path: Path):
        writer = ReconciliationExcelWriter(special_conditions_text="Особый порядок уведомления")
        output_path = tmp_path / "out.xlsx"
        writer.write(_sample_report(), output_path)
        assert "Особые условия" in load_workbook(output_path).sheetnames

    def test_no_special_conditions_sheet_when_empty(self, tmp_path: Path):
        writer = ReconciliationExcelWriter(special_conditions_text="")
        output_path = tmp_path / "out.xlsx"
        writer.write(_sample_report(), output_path)
        assert "Особые условия" not in load_workbook(output_path).sheetnames


class TestRealRepoTemplates:
    def test_both_templates_bundled(self):
        assert VERTICAL_TEMPLATE.exists()
        assert HORIZONTAL_TEMPLATE.exists()

    def test_writes_using_repo_template_headers(self, tmp_path: Path):
        template = ResponseTemplate.load(VERTICAL_TEMPLATE)
        writer = ReconciliationExcelWriter(template=template)
        output_path = tmp_path / "200 – результат проверки.xlsx"

        writer.write(_sample_report(), output_path)

        ws = load_workbook(output_path).active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert headers[0] == "Номер декларации/строки"
        assert "совпадает" in headers[3]

    def test_output_follows_template_field_list(self, tmp_path: Path):
        """The form prescribes the fields — output rows must follow it, not
        just echo whatever the model happened to return."""
        template = ResponseTemplate.load(VERTICAL_TEMPLATE)
        writer = ReconciliationExcelWriter(template=template)
        output_path = tmp_path / "out.xlsx"

        writer.write(_sample_report(), output_path)

        ws = load_workbook(output_path).active
        written_fields = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        for field in template.fields:
            assert field.field_name in written_fields


class TestTemplateFieldSkeleton:
    def test_missing_model_row_is_marked_not_checked(self, tmp_path: Path):
        template = _stub_template(tmp_path, [
            TemplateField("Объект страхования", "Объект страхования"),
            TemplateField("Пломбирование", "НЕ ОБЯЗАТЕЛЕН ДЛЯ СВЕРКИ"),
        ])
        report = ReconciliationReport(
            declaration_number="200",
            rows=[ReconciliationRow("200", "Объект страхования", "3.2", "совпадает", "ок")],
        )

        ReconciliationExcelWriter(template=template).write(report, tmp_path / "out.xlsx")

        ws = load_workbook(tmp_path / "out.xlsx").active
        rows = {ws.cell(row=r, column=2).value: ws.cell(row=r, column=4).value
                for r in range(2, ws.max_row + 1)}
        assert rows["Объект страхования"] == "совпадает"
        assert rows["Пломбирование"] == "не проверено"

    def test_template_wording_wins_over_model_paraphrase(self, tmp_path: Path):
        template = _stub_template(tmp_path, [TemplateField("Полное наименование груза", "Объект страхования")])
        report = ReconciliationReport(
            declaration_number="200",
            # model paraphrased the field name
            rows=[ReconciliationRow("200", "наименование груза", "п.9", "совпадает", "ок")],
        )

        ReconciliationExcelWriter(template=template).write(report, tmp_path / "out.xlsx")

        ws = load_workbook(tmp_path / "out.xlsx").active
        assert ws.cell(row=2, column=2).value == "Полное наименование груза"
        assert ws.cell(row=2, column=3).value == "Объект страхования"

    def test_extra_model_rows_are_kept(self, tmp_path: Path):
        template = _stub_template(tmp_path, [TemplateField("Объект страхования", "Объект страхования")])
        report = ReconciliationReport(
            declaration_number="200",
            rows=[
                ReconciliationRow("200", "Объект страхования", "9", "совпадает", "ок"),
                ReconciliationRow("200", "Неожиданное поле", "—", "не совпадает", "находка"),
            ],
        )

        ReconciliationExcelWriter(template=template).write(report, tmp_path / "out.xlsx")

        ws = load_workbook(tmp_path / "out.xlsx").active
        written = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "Неожиданное поле" in written

    def test_multi_shipment_repeats_field_block_per_line_item(self, tmp_path: Path):
        template = _stub_template(tmp_path, [
            TemplateField("Объект страхования", "Объект страхования"),
            TemplateField("Маршрут", "Маршрут перевозок"),
        ])
        report = ReconciliationReport(
            declaration_number="200",
            rows=[
                ReconciliationRow("200/1", "Объект страхования", "9", "совпадает", "ок"),
                ReconciliationRow("200/2", "Объект страхования", "9", "не совпадает", "иначе"),
            ],
        )

        ReconciliationExcelWriter(template=template).write(report, tmp_path / "out.xlsx")

        ws = load_workbook(tmp_path / "out.xlsx").active
        refs = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert refs.count("200/1") == 2   # both prescribed fields, per line item
        assert refs.count("200/2") == 2

    def test_contradictory_match_is_flagged_in_comment(self, tmp_path: Path):
        template = _stub_template(tmp_path, [TemplateField("Тариф", "Страховой тариф")])
        report = ReconciliationReport(
            declaration_number="200",
            rows=[ReconciliationRow(
                "200", "Тариф", "п.5", "совпадает",
                "В декларации 0,15%, а в полисе 0,2% — значения отличаются",
                needs_review=True,
            )],
        )

        ReconciliationExcelWriter(template=template).write(report, tmp_path / "out.xlsx")

        ws = load_workbook(tmp_path / "out.xlsx").active
        assert "ТРЕБУЕТ ПРОВЕРКИ" in ws.cell(row=2, column=5).value
