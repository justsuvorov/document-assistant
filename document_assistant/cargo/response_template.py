from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from document_assistant.cargo.declaration_classifier import DeclarationType
from document_assistant.core.settings import settings

_BUNDLED_DIR = Path(__file__).parent / "templates"
_BUNDLED_HORIZONTAL = _BUNDLED_DIR / "reconciliation_form_horizontal.xlsx"
_BUNDLED_VERTICAL = _BUNDLED_DIR / "reconciliation_form_vertical.xlsx"


@dataclass(frozen=True)
class TemplateField:
    """One prescribed row of the response form: which declaration field to
    check, and which policy clause it is checked against."""
    field_name: str
    policy_clause: str


@dataclass
class ResponseTemplate:
    """The response form («Структура проверки декларации») for one declaration
    layout.

    The form is not just headers — rows 2..N prescribe the EXACT field list to
    check and the policy clause each field maps to. That list drives both ends
    of the pipeline: it is fed to the LLM so it reports on exactly these fields
    (in this order), and it is the row skeleton the writer fills in.
    """
    path: Path
    fields: list[TemplateField]

    HEADER_ROW = 1
    FIELD_NAME_COL = 2      # B — «Наименование поля в декларации»
    POLICY_CLAUSE_COL = 3   # C — «С каким пунктом Ген. полиса сверено»

    @classmethod
    def load(cls, path: str | Path) -> "ResponseTemplate":
        path = Path(path)
        wb = load_workbook(path)
        ws = wb.active

        fields = []
        for row in range(cls.HEADER_ROW + 1, ws.max_row + 1):
            field_name = ws.cell(row=row, column=cls.FIELD_NAME_COL).value
            policy_clause = ws.cell(row=row, column=cls.POLICY_CLAUSE_COL).value
            if field_name is None or not str(field_name).strip():
                continue
            fields.append(TemplateField(
                field_name=str(field_name).strip(),
                policy_clause=str(policy_clause).strip() if policy_clause else "",
            ))
        wb.close()

        print(f"[INFO] Форма ответа: {path.name} — {len(fields)} полей для проверки", flush=True)
        return cls(path=path, fields=fields)

    def to_prompt_block(self) -> str:
        """Markdown table of the prescribed fields, injected into the
        reconciliation prompt so the LLM answers for exactly these rows."""
        if not self.fields:
            return ""
        lines = ["| Наименование поля в декларации | С каким пунктом Ген. полиса сверено |", "|---|---|"]
        for f in self.fields:
            clause = f.policy_clause.replace("\n", " ")
            lines.append(f"| {f.field_name} | {clause} |")
        return "\n".join(lines)


class ResponseTemplateResolver:
    """Picks the response form matching the declaration's layout.

    Horizontal (ПСГ) form — a «№ п/п» table with one row per shipment, i.e.
    what DeclarationTypeClassifier reports as MULTI.
    Vertical form — a top-to-bottom field list for a single shipment (SINGLE).
    """

    @staticmethod
    def path_for(decl_type: DeclarationType) -> Path:
        if decl_type is DeclarationType.MULTI:
            configured = settings.reconciliation_template_horizontal
            bundled = _BUNDLED_HORIZONTAL
        else:
            configured = settings.reconciliation_template_vertical
            bundled = _BUNDLED_VERTICAL

        if configured and Path(configured).exists():
            return Path(configured)
        if configured:
            print(
                f"[WARN] Форма ответа: путь из .env не найден ({configured}), "
                f"используется встроенная — {bundled.name}",
                flush=True,
            )
        return bundled

    @classmethod
    def for_type(cls, decl_type: DeclarationType) -> ResponseTemplate:
        return ResponseTemplate.load(cls.path_for(decl_type))
