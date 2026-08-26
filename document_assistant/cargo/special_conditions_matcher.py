import re

import openpyxl


class PolicyIdentityExtractor:
    """Extracts the policy number and insurant name from the general policy
    document's own text, so special-conditions lookup can find "which policy
    is this" without any manual input.

    Policy numbers follow VSK's internal scheme (e.g. "2518013GR1941",
    "2600A13GR1339AZOAT", "2506Q13GR2182") — every observed sample contains
    the substring "13G", which is distinctive enough to find the number
    without requiring a specific "№"/label context that varies by document.
    """

    _NUMBER_RE = re.compile(r"\b([0-9A-ZА-Я]{4,20}13G[0-9A-ZА-Я]{2,14})\b", re.IGNORECASE)
    _INSURANT_RE = re.compile(r"Страхователь[:\s|]+([^\n|]{3,150})", re.IGNORECASE)

    def extract_number(self, policy_text: str) -> str | None:
        match = self._NUMBER_RE.search(policy_text)
        return match.group(1).upper() if match else None

    def extract_insurant(self, policy_text: str) -> str | None:
        match = self._INSURANT_RE.search(policy_text)
        return match.group(1).strip() if match else None


class SpecialConditionsMatcher:
    """Looks up the current policy's special conditions inside the shared
    "ДпГ УПиСД ФСЦ" workbook (SPECIAL_CONDITIONS_GLOBAL_PATH) instead of
    dumping the whole multi-thousand-row, multi-sheet file into the prompt —
    which used to blow the context window and cause 400 Bad Request.

    Only two sheets matter, matched by policy identity:
        "ГП"        — one row per general policy. Matched by policy number
                      (column "№ Генерального полиса"); the "Особые условия"
                      cell of the matching row is the payload.
        "ОБОРОННЫЙ" — defense-sector clients requiring reinsurance on every
                      declaration. Matched by contract number (column
                      "Договор") or insurant name (column "Страхователь");
                      presence of a matching row IS the special condition
                      (there's no separate free-text column here).
    """

    SHEET_GP = "ГП"
    SHEET_DEFENSE = "ОБОРОННЫЙ"

    _COL_GP_NUMBER = "№ Генерального полиса"
    _COL_GP_INSURANT = "Наименование Страхователя"
    _COL_GP_CONDITIONS = "Особые условия"

    _COL_DEF_INSURANT = "Страхователь"
    _COL_DEF_CONTRACT = "Договор"
    _COL_DEF_REINSURANCE = "Перестрахование, при условии:"
    _COL_DEF_FORM = "Форма работы"

    def match(self, xlsx_path: str, policy_number: str | None, insurant: str | None) -> str:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        parts = []

        if self.SHEET_GP in wb.sheetnames:
            gp_text = self._match_gp_sheet(wb[self.SHEET_GP], policy_number)
            if gp_text:
                parts.append(gp_text)

        if self.SHEET_DEFENSE in wb.sheetnames:
            defense_text = self._match_defense_sheet(wb[self.SHEET_DEFENSE], policy_number, insurant)
            if defense_text:
                parts.append(defense_text)

        return "\n\n".join(parts)

    def _match_gp_sheet(self, ws, policy_number: str | None) -> str:
        if not policy_number:
            return ""

        header = self._header_index(ws)
        num_col = header.get(self._norm(self._COL_GP_NUMBER))
        cond_col = header.get(self._norm(self._COL_GP_CONDITIONS))
        if num_col is None or cond_col is None:
            print(
                f"[WARN] Лист «{self.SHEET_GP}» в файле особых условий: не найдены колонки "
                f"«{self._COL_GP_NUMBER}» / «{self._COL_GP_CONDITIONS}»",
                flush=True,
            )
            return ""

        found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_num = row[num_col] if num_col < len(row) else None
            if cell_num is None or self._norm_number(str(cell_num)) != policy_number:
                continue
            condition = row[cond_col] if cond_col < len(row) else None
            if condition:
                found.append(str(condition).strip())

        if not found:
            return ""
        return "### Особые условия по ГП (найдено по № " + policy_number + ")\n\n" + "\n".join(
            f"- {c}" for c in found
        )

    def _match_defense_sheet(self, ws, policy_number: str | None, insurant: str | None) -> str:
        header = self._header_index(ws)
        insurant_col = header.get(self._norm(self._COL_DEF_INSURANT))
        contract_col = header.get(self._norm(self._COL_DEF_CONTRACT))
        reinsurance_col = header.get(self._norm(self._COL_DEF_REINSURANCE))
        form_col = header.get(self._norm(self._COL_DEF_FORM))
        if insurant_col is None and contract_col is None:
            print(
                f"[WARN] Лист «{self.SHEET_DEFENSE}» в файле особых условий: не найдены колонки "
                f"«{self._COL_DEF_INSURANT}» / «{self._COL_DEF_CONTRACT}»",
                flush=True,
            )
            return ""

        insurant_norm = self._norm_text(insurant) if insurant else None
        found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            contract_val = row[contract_col] if contract_col is not None and contract_col < len(row) else None
            insurant_val = row[insurant_col] if insurant_col is not None and insurant_col < len(row) else None

            contract_match = (
                policy_number is not None and contract_val is not None
                and self._norm_number(str(contract_val)) == policy_number
            )
            insurant_match = (
                insurant_norm is not None and insurant_val is not None
                and self._norm_text(str(insurant_val)) == insurant_norm
            )
            if not (contract_match or insurant_match):
                continue

            reinsurance = row[reinsurance_col] if reinsurance_col is not None and reinsurance_col < len(row) else ""
            form = row[form_col] if form_col is not None and form_col < len(row) else ""
            label = str(insurant_val) if insurant_val else (str(contract_val) if contract_val else "")
            found.append(
                f"- {label}: перестрахование при условии «{reinsurance}», форма работы «{form}»"
            )

        if not found:
            return ""
        return f"### Клиент относится к оборонному сегменту (лист «{self.SHEET_DEFENSE}»)\n\n" + "\n".join(found)

    @staticmethod
    def _header_index(ws) -> dict[str, int]:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return {
            SpecialConditionsMatcher._norm(str(cell)): idx
            for idx, cell in enumerate(header_row)
            if cell
        }

    @staticmethod
    def _norm(text: str) -> str:
        return " ".join(text.lower().split())

    @staticmethod
    def _norm_number(text: str) -> str:
        return re.sub(r"\s+", "", text).upper()

    @staticmethod
    def _norm_text(text: str) -> str:
        cleaned = re.sub(r"[«»\"'«»]", "", text)
        return " ".join(cleaned.lower().split())
