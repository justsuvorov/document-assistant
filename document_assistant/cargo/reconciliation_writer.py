import re
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from document_assistant.cargo.models import ReconciliationReport, ReconciliationRow
from document_assistant.cargo.response_template import ResponseTemplate
from document_assistant.cargo.result_consistency import MATCH
from document_assistant.reports.style import GREEN, RED, YELLOW
from document_assistant.reports.style import status_fill as _shared_status_fill
from document_assistant.reports.writers import ReportWriter

_STATUS_FILL = {
    "совпадает":     GREEN,
    "не совпадает":  RED,
}
_REVIEW_FILL = YELLOW


def _status_fill(status: str) -> str:
    return _shared_status_fill(status, _STATUS_FILL)


class ReconciliationExcelWriter(ReportWriter):
    """Writes a ReconciliationReport into a fresh copy of the response form
    («Структура проверки декларации»).

    The form prescribes the field list, so this writer does not simply dump
    whatever rows the model returned: it walks the template's fields in order
    and fills each one from the matching model row, leaving an explicit
    "не проверено" marker where the model returned nothing. That keeps every
    output file structurally identical to the form the business signed off on,
    regardless of what the model chose to answer.

    For a multi-shipment declaration the field block repeats once per line
    item (200/1, 200/2, …), all inside the one output file.
    """

    HEADERS = [
        "Номер декларации/строки",
        "Наименование поля в декларации",
        "С каким пунктом Ген. полиса сверено",
        "Результат проверки\n(совпадает/не совпадает)",
        "Комментарий по сверке",
    ]

    _NOT_CHECKED = "не проверено"

    def __init__(self, template: ResponseTemplate | None = None, special_conditions_text: str = ""):
        self._template = template
        self._special_conditions_text = special_conditions_text

    def write(self, report: ReconciliationReport, output_path: Path, source_path: Path = None) -> Path:
        if self._template is not None and self._template.path.exists():
            shutil.copy2(self._template.path, output_path)
            wb = load_workbook(output_path)
            ws = wb.active
            self._clear_template_body(ws)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Лист1"
            for col, title in enumerate(self.HEADERS, start=1):
                ws.cell(row=1, column=col, value=title)

        self._write_rows(ws, report)
        self._autofit_column_widths(ws)
        ws.freeze_panes = "A2"

        if self._special_conditions_text:
            self._write_notes_sheet(wb, self._special_conditions_text)

        wb.save(output_path)
        return output_path

    @staticmethod
    def _clear_template_body(ws) -> None:
        """The template ships with its field list pre-filled as example rows;
        results are written back over them, so start from the header only."""
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    def _rows_to_write(self, report: ReconciliationReport) -> list[ReconciliationRow]:
        if self._template is None or not self._template.fields:
            return report.rows

        by_ref: dict[str, list[ReconciliationRow]] = {}
        for row in report.rows:
            by_ref.setdefault(row.declaration_ref, []).append(row)
        if not by_ref:
            by_ref = {report.declaration_number: []}

        ordered: list[ReconciliationRow] = []
        for ref, model_rows in by_ref.items():
            remaining = list(model_rows)
            for field in self._template.fields:
                match = self._pop_match(remaining, field.field_name)
                if match is not None:
                    # Keep the form's own wording/clause mapping — the model
                    # paraphrases them, and the business reads this column.
                    match.field_name = field.field_name
                    match.matched_policy_clause = field.policy_clause or match.matched_policy_clause
                    ordered.append(match)
                else:
                    ordered.append(ReconciliationRow(
                        declaration_ref=ref,
                        field_name=field.field_name,
                        matched_policy_clause=field.policy_clause,
                        result=self._NOT_CHECKED,
                        comment="Модель не вернула результат по этому полю",
                        needs_review=True,
                    ))
            # Anything the model reported beyond the form's field list is kept
            # rather than dropped — it may be a genuine finding.
            ordered.extend(remaining)
        return ordered

    @classmethod
    def _pop_match(cls, rows: list[ReconciliationRow], field_name: str) -> ReconciliationRow | None:
        target = cls._normalize(field_name)
        for i, row in enumerate(rows):
            if cls._normalize(row.field_name) == target:
                return rows.pop(i)
        target_words = cls._words(field_name)
        if not target_words:
            return None
        best_i, best_score = None, 0.0
        for i, row in enumerate(rows):
            row_words = cls._words(row.field_name)
            if not row_words:
                continue
            overlap = len(target_words & row_words) / min(len(target_words), len(row_words))
            if overlap > best_score:
                best_i, best_score = i, overlap
        return rows.pop(best_i) if best_i is not None and best_score >= 0.6 else None

    @staticmethod
    def _normalize(text: str) -> str:
        return " ".join((text or "").lower().replace("\n", " ").split())

    @staticmethod
    def _words(text: str) -> set[str]:
        return {w for w in re.split(r"\W+", (text or "").lower()) if len(w) > 2}

    def _write_rows(self, ws, report: ReconciliationReport) -> None:
        wrap = Alignment(vertical="top", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_idx = 2
        for row in self._rows_to_write(report):
            comment = row.comment
            if row.needs_review and row.result == MATCH:
                comment = f"[ТРЕБУЕТ ПРОВЕРКИ: комментарий указывает на расхождение] {comment}"

            values = [row.declaration_ref, row.field_name, row.matched_policy_clause, row.result, comment]
            fill_color = _REVIEW_FILL if row.needs_review else _status_fill(row.result)
            fill = PatternFill("solid", fgColor=fill_color)
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = wrap
                cell.border = border
                if col == 4:
                    cell.fill = fill
            row_idx += 1

    @staticmethod
    def _autofit_column_widths(ws, min_width: int = 10, max_width: int = 60) -> None:
        """Size each column to its content (longest line per cell), capped so a
        single verbose comment can't blow the whole column out — long text
        still wraps thanks to the ``wrap_text`` alignment set in _write_rows."""
        widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                longest_line = max((len(line) for line in str(cell.value).split("\n")), default=0)
                widths[cell.column] = max(widths.get(cell.column, 0), longest_line)

        for col, length in widths.items():
            width = max(min_width, min(length + 2, max_width))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    @staticmethod
    def _write_notes_sheet(wb, text: str) -> None:
        ws = wb.create_sheet(title="Особые условия")
        ws.column_dimensions["A"].width = 100
        cell = ws.cell(row=1, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
